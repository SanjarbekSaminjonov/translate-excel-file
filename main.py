from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from translator import latin2cyrillic, cyrillic2latin


class Excel:
    def __init__(self, file_name: str, to_latin: bool = False):
        self._file_name = file_name
        self._convertor_func = cyrillic2latin if to_latin else latin2cyrillic
        try:
            self._wb = load_workbook(self._file_name)
        except FileNotFoundError:
            self._wb = Workbook()
        self._ws = self._wb.active

    def get_sheets_list(self):
        return self._wb.sheetnames

    def set_sheet(self, sheet_name: str):
        try:
            self._ws = self._wb[sheet_name]
        except KeyError:
            self._ws = self._wb.create_sheet(sheet_name)

    def translate_sheet(self):
        for row in self._ws.rows:
            for cell in row:
                if type(cell.value) == str:
                    cell.value = self._convertor_func(cell.value)
        self._ws.title = self._convertor_func(self._ws.title)

    def translate_all_sheets(self):
        for sheet_name in self.get_sheets_list():
            self.set_sheet(sheet_name)
            self.translate_sheet()

    def save(self, new_file_name=None):
        if new_file_name:
            self._wb.save(new_file_name)
        else:
            self._wb.save(self._file_name)


# create excel obj to translate
ex = Excel('latin.xlsx', to_latin=False)

# translate sheets one by one
# for sheet in ex.get_sheets_list():
#     ex.set_sheet(sheet)
#     ex.translate_sheet()

# translate all sheets at once
ex.translate_all_sheets()

# save file with new file name or old name (to save with old name use ex.save())
ex.save('cyrillic.xlsx')
